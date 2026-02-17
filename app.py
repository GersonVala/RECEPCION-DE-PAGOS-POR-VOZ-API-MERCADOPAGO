from datetime import datetime, timezone
import random
import re
import threading
import time

from io import BytesIO

from functools import wraps
from flask import Flask, request, render_template, jsonify, send_file, session, redirect, url_for
import requests

from config import MP_ACCESS_TOKEN, FLASK_PORT
from database import init_db, insert_payment, get_payments, get_totals, get_payments_by_period
from tts import announce_payment

app = Flask(__name__)
app.secret_key = "holagranja-secret-key-2026"

DASHBOARD_PASSWORD = "miguel2026"


def login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if not session.get("authenticated"):
            return redirect(url_for("login"))
        return f(*args, **kwargs)
    return decorated

# Datos de nuestra cuenta MP (se obtienen al iniciar)
MY_USER_ID = None
MY_USER_NAME = ""
MY_USER_EMAIL = ""


def fetch_my_user_info():
    """Obtiene el user_id, nombre y email de nuestra cuenta MP."""
    headers = {"Authorization": f"Bearer {MP_ACCESS_TOKEN}"}
    try:
        response = requests.get("https://api.mercadopago.com/users/me", headers=headers, timeout=10)
        if response.status_code == 200:
            data = response.json()
            uid = data.get("id")
            fname = (data.get("first_name", "") or "").strip()
            lname = (data.get("last_name", "") or "").strip()
            full_name = f"{fname} {lname}".strip()
            email = (data.get("email", "") or "").strip().lower()
            print(f"[MP] Cuenta identificada - User ID: {uid}, Nombre: {full_name}, Email: {email}")
            return uid, full_name, email
    except requests.RequestException:
        pass
    return None, "", ""


# --- Polling: consulta la API de MP cada 15 segundos como respaldo del webhook ---

def process_payment_info(payment_info):
    """Procesa un pago obtenido de la API de MP (usado por webhook y polling)."""
    # Ignorar pagos salientes (transferencias que nosotros enviamos)
    operation_type = payment_info.get("operation_type", "")
    if operation_type in ("money_transfer", "account_fund"):
        # Verificar si es un pago saliente comparando collector con nuestra cuenta
        collector = payment_info.get("collector_id") or (payment_info.get("collector", {}) or {}).get("id")
        if MY_USER_ID and collector and str(collector) != str(MY_USER_ID):
            return False

    payer = payment_info.get("payer", {})
    payer_id = payer.get("id")
    payer_email = (payer.get("email", "") or "").strip()

    first_name = (payer.get("first_name", "") or "").strip()
    last_name = (payer.get("last_name", "") or "").strip()
    payer_name = f"{first_name} {last_name}".strip()

    # Para transferencias (account_money/account_fund), MP devuelve datos del
    # COBRADOR en los campos payer en vez del pagador real.
    # Detectamos esto por ID, nombre o email coincidente con nuestra cuenta.
    id_matches = MY_USER_ID and payer_id and str(payer_id) == str(MY_USER_ID)
    name_matches = MY_USER_NAME and payer_name and payer_name.lower() == MY_USER_NAME.lower()
    email_matches = MY_USER_EMAIL and payer_email and payer_email.lower() == MY_USER_EMAIL.lower()
    datos_son_nuestros = id_matches or name_matches or email_matches

    if datos_son_nuestros:
        payer_name = ""
        payer_email = ""

        # Fuente 1: bank_info.payer.long_name (transferencias bancarias/Personal Pay)
        poi = payment_info.get("point_of_interaction") or {}
        td = poi.get("transaction_data") or {}
        bank_payer = (td.get("bank_info") or {}).get("payer") or {}
        long_name = (bank_payer.get("long_name", "") or "").strip()
        if long_name and long_name.lower() != MY_USER_NAME.lower():
            payer_name = long_name

        # Fuente 2: additional_info.payer
        if not payer_name:
            alt_payer = (payment_info.get("additional_info") or {}).get("payer") or {}
            alt_first = (alt_payer.get("first_name", "") or "").strip()
            alt_last = (alt_payer.get("last_name", "") or "").strip()
            alt_name = f"{alt_first} {alt_last}".strip()
            if alt_name and alt_name.lower() != MY_USER_NAME.lower():
                payer_name = alt_name

        # Fuente 3: email del payer (solo si NO es nuestro email)
        raw_email = (payer.get("email", "") or "").strip()
        if raw_email and raw_email.lower() != MY_USER_EMAIL.lower():
            payer_email = raw_email
            # Derivar nombre del email solo si es de otra persona
            if not payer_name:
                raw = raw_email.split("@")[0].replace("_", " ").replace(".", " ")
                clean = re.sub(r'[^a-zA-ZáéíóúñÁÉÍÓÚÑ ]', '', raw).strip()
                if clean:
                    payer_name = clean.title()

        payer_name = payer_name or "Cliente"
    else:
        # Datos normales (tarjeta, QR, etc.) - confiar en payer
        if not payer_name and payer_email:
            raw = payer_email.split("@")[0].replace("_", " ").replace(".", " ")
            clean = re.sub(r'[^a-zA-ZáéíóúñÁÉÍÓÚÑ ]', '', raw).strip()
            if clean:
                payer_name = clean.title()

        payer_name = payer_name or "Cliente"

    payment_type_id = payment_info.get("payment_type_id", "")
    type_map = {
        "account_money": "Transferencia",
        "bank_transfer": "Transferencia",
        "credit_card": "Tarjeta de credito",
        "debit_card": "Tarjeta de debito",
        "prepaid_card": "Tarjeta prepaga",
    }
    payment_type = type_map.get(payment_type_id, "Transferencia")

    payment_data = {
        "mp_payment_id": payment_info.get("id"),
        "payer_name": payer_name,
        "payer_email": payer_email,
        "amount": payment_info.get("transaction_amount", 0),
        "status": payment_info.get("status", ""),
        "payment_type": payment_type,
        "date_created": payment_info.get("date_created", ""),
    }

    inserted = insert_payment(payment_data)

    if inserted and payment_data["status"] in ("approved", "rejected"):
        say_name = payer_name if payer_name not in ("Cliente", "Transferencia Recibida") else None
        announce_payment(say_name, payment_data["amount"], rejected=(payment_data["status"] == "rejected"))

    return inserted


def poll_payments():
    """Hilo de conciliacion: consulta pagos recientes a la API de MP cada 60s como backup del webhook."""
    headers = {"Authorization": f"Bearer {MP_ACCESS_TOKEN}"}
    last_check = datetime.now(timezone.utc)

    while True:
        try:
            now = datetime.now(timezone.utc)
            params = {
                "sort": "date_created",
                "criteria": "desc",
                "begin_date": last_check.strftime("%Y-%m-%dT%H:%M:%SZ"),
                "end_date": now.strftime("%Y-%m-%dT%H:%M:%SZ"),
                "range": "date_created",
                "status": "approved",
            }
            response = requests.get(
                "https://api.mercadopago.com/v1/payments/search",
                headers=headers, params=params, timeout=15
            )
            if response.status_code == 200:
                results = response.json().get("results", [])
                for result in results:
                    pid = result.get("id")
                    # Consultar detalles completos (misma ruta que el webhook)
                    payment_info = fetch_payment_details(pid)
                    if not payment_info:
                        continue
                    inserted = process_payment_info(payment_info)
                    if inserted:
                        print(f"[Polling] Pago detectado - ID: {pid}")
            elif response.status_code == 429:
                print("[Polling] Rate limit, esperando 60s...")
                time.sleep(60)

            last_check = now
        except Exception as e:
            print(f"[Polling] Error: {e}")

        time.sleep(60)


def _process_webhook_payment(payment_id):
    """Procesa un pago del webhook en hilo separado."""
    payment_info = fetch_payment_details(payment_id)
    if payment_info:
        process_payment_info(payment_info)


@app.route("/webhook", methods=["POST", "GET"])
def webhook():
    # Formato IPN: MP envia topic e id como query params
    topic_param = request.args.get("topic", "")
    id_param = request.args.get("id", "")

    # Formato Webhook v2: MP envia JSON en el body
    body = request.get_json(silent=True) or {}
    topic_body = body.get("type", body.get("topic", ""))
    id_body = (body.get("data") or {}).get("id", "")

    # Usar el que tenga datos
    topic = topic_param or topic_body
    payment_id = id_param or id_body

    # Solo procesar notificaciones de pago
    if "payment" not in topic:
        return "OK", 200

    if not payment_id:
        return "OK", 200

    print(f"[Webhook] Pago recibido - ID: {payment_id}")

    # Responder 200 inmediatamente y procesar en hilo aparte
    # MP espera respuesta rapida, si no reintenta innecesariamente
    threading.Thread(target=_process_webhook_payment, args=(payment_id,), daemon=True).start()

    return "OK", 200



def fetch_payment_details(payment_id):
    url = f"https://api.mercadopago.com/v1/payments/{payment_id}"
    headers = {"Authorization": f"Bearer {MP_ACCESS_TOKEN}"}
    for attempt in range(3):
        try:
            response = requests.get(url, headers=headers, timeout=15)
            if response.status_code == 200:
                return response.json()
            # Errores 5xx son transitorios, reintentar
            if response.status_code >= 500:
                print(f"[MP API] Error {response.status_code} (intento {attempt + 1}/3)")
                time.sleep(2)
                continue
            # Errores 4xx no se reintentan
            print(f"[MP API] Error {response.status_code}: {response.text}")
            return None
        except requests.RequestException as e:
            print(f"[MP API] Intento {attempt + 1}/3 fallo: {e}")
            time.sleep(2)
    return None


@app.route("/login", methods=["GET", "POST"])
def login():
    error = ""
    if request.method == "POST":
        if request.form.get("password") == DASHBOARD_PASSWORD:
            session["authenticated"] = True
            return redirect(url_for("index"))
        error = "Contraseña incorrecta"
    return render_template("login.html", error=error)


@app.route("/logout")
def logout():
    session.clear()
    return redirect(url_for("login"))


@app.route("/")
@login_required
def index():
    date_from = request.args.get("fecha_desde", "").strip()
    date_to = request.args.get("fecha_hasta", "").strip()
    amount_min = request.args.get("monto_min", "").strip()
    amount_max = request.args.get("monto_max", "").strip()
    page = max(1, int(request.args.get("page", 1)))

    payments, total, total_pages = get_payments(
        date_from=date_from or None,
        date_to=date_to or None,
        amount_min=amount_min or None,
        amount_max=amount_max or None,
        page=page,
    )

    totals = get_totals()

    return render_template(
        "index.html",
        payments=payments,
        total=total,
        total_pages=total_pages,
        page=page,
        totals=totals,
        filters={
            "fecha_desde": date_from,
            "fecha_hasta": date_to,
            "monto_min": amount_min,
            "monto_max": amount_max,
        },
    )


@app.route("/test")
@login_required
def test_payment():
    """Simula un pago para probar el sistema sin Mercado Pago."""
    names = ["Juan Perez", "Maria Garcia", "Carlos Lopez", "Ana Martinez", "Pedro Sanchez"]
    name = random.choice(names)
    amount = random.choice([500, 1200, 3500, 7800, 15000])

    payment_data = {
        "mp_payment_id": f"TEST_{random.randint(100000, 999999)}",
        "payer_name": name,
        "payer_email": f"{name.split()[0].lower()}@test.com",
        "amount": amount,
        "description": "Pago de prueba",
        "status": "approved",
        "payment_type": "Transferencia",
        "date_created": datetime.now().isoformat(),
    }

    insert_payment(payment_data)
    announce_payment(name, amount)

    return f"Pago simulado: {name} - ${amount}", 200


@app.route("/debug/payment/<payment_id>")
@login_required
def debug_payment(payment_id):
    """Muestra el JSON crudo que devuelve la API de MP para un pago."""
    import json
    payment_info = fetch_payment_details(payment_id)
    if not payment_info:
        return jsonify({"error": "No se pudo obtener el pago"}), 404

    # Extraer los campos relevantes para el diagnostico
    payer = payment_info.get("payer", {})
    collector = payment_info.get("collector", {})
    poi = payment_info.get("point_of_interaction", {})
    td = (poi.get("transaction_data") or {})
    bank_info = td.get("bank_info") or {}
    additional = payment_info.get("additional_info") or {}

    debug = {
        "_DIAGNOSTICO": "Campos relevantes para identificar al pagador real",
        "1_PAYER_toplevel": {
            "id": payer.get("id"),
            "first_name": payer.get("first_name"),
            "last_name": payer.get("last_name"),
            "email": payer.get("email"),
            "identification": payer.get("identification"),
            "operator_id": payer.get("operator_id"),
        },
        "2_COLLECTOR": {
            "id": payment_info.get("collector_id"),
            "collector_obj": collector,
        },
        "3_BANK_INFO_PAYER": bank_info.get("payer"),
        "4_BANK_INFO_COLLECTOR": bank_info.get("collector"),
        "5_BANK_INFO_extras": {
            "origin_bank_id": bank_info.get("origin_bank_id"),
            "origin_wallet_id": bank_info.get("origin_wallet_id"),
            "is_same_bank_account_owner": bank_info.get("is_same_bank_account_owner"),
        },
        "6_POINT_OF_INTERACTION": {
            "type": poi.get("type"),
            "sub_type": poi.get("sub_type"),
            "transaction_data_e2e_id": td.get("e2e_id"),
            "transaction_data_transaction_id": td.get("transaction_id"),
        },
        "7_ADDITIONAL_INFO_PAYER": additional.get("payer"),
        "8_METADATA": payment_info.get("metadata"),
        "9_DESCRIPTION": payment_info.get("description"),
        "10_OPERATION_TYPE": payment_info.get("operation_type"),
        "11_PAYMENT_METHOD": payment_info.get("payment_method_id"),
        "12_PAYMENT_TYPE": payment_info.get("payment_type_id"),
        "RAW_COMPLETO": payment_info,
    }
    return app.response_class(
        response=json.dumps(debug, indent=2, ensure_ascii=False, default=str),
        status=200,
        mimetype="application/json",
    )


@app.route("/api/pagos")
@login_required
def api_pagos():
    page = max(1, int(request.args.get("page", 1)))
    payments, total, total_pages = get_payments(
        date_from=request.args.get("fecha_desde") or None,
        date_to=request.args.get("fecha_hasta") or None,
        amount_min=request.args.get("monto_min") or None,
        amount_max=request.args.get("monto_max") or None,
        page=page,
    )
    totals = get_totals(
        dia=request.args.get("totals_dia") or None,
        mes=request.args.get("totals_mes") or None,
        anio=request.args.get("totals_anio") or None,
    )
    return jsonify({"payments": payments, "totals": totals, "page": page, "total_pages": total_pages, "total": total})


@app.route("/api/exportar")
@login_required
def exportar_excel():
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

    periodo = request.args.get("periodo", "dia")
    valor = request.args.get("valor", "")

    if not valor:
        return "Falta el parametro 'valor'", 400

    payments = get_payments_by_period(periodo, valor)

    # Mapeo de tipos para registros viejos
    type_map = {
        "bank_transfer": "Transferencia",
        "account_money": "Transferencia",
        "credit_card": "Tarjeta de credito",
        "debit_card": "Tarjeta de debito",
        "prepaid_card": "Tarjeta prepaga",
    }

    wb = Workbook()
    ws = wb.active

    if periodo == "dia":
        ws.title = f"Ventas {valor}"
        titulo = f"Ventas del dia {valor}"
    elif periodo == "mes":
        ws.title = f"Ventas {valor}"
        titulo = f"Ventas del mes {valor}"
    else:
        ws.title = f"Ventas {valor}"
        titulo = f"Ventas del año {valor}"

    # Estilos
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="343A40", end_color="343A40", fill_type="solid")
    title_font = Font(bold=True, size=14)
    total_font = Font(bold=True, size=12)
    border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # Titulo
    ws.merge_cells("A1:E1")
    ws["A1"] = titulo
    ws["A1"].font = title_font
    ws["A1"].alignment = Alignment(horizontal="center")

    # Headers
    headers = ["Fecha", "Nombre", "Email", "Tipo de pago", "Monto"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = border

    # Datos
    total = 0
    for i, p in enumerate(payments, 4):
        fecha = (p.get("date_created") or "-")[:19]
        tipo = type_map.get(p.get("payment_type", ""), p.get("payment_type", "-"))
        monto = p.get("amount", 0)
        total += monto

        ws.cell(row=i, column=1, value=fecha).border = border
        ws.cell(row=i, column=2, value=p.get("payer_name", "")).border = border
        ws.cell(row=i, column=3, value=p.get("payer_email", "")).border = border
        ws.cell(row=i, column=4, value=tipo).border = border
        cell_monto = ws.cell(row=i, column=5, value=monto)
        cell_monto.number_format = '#,##0.00'
        cell_monto.alignment = Alignment(horizontal="right")
        cell_monto.border = border

    # Fila de total
    total_row = len(payments) + 4
    total_fill = PatternFill(start_color="D4EDDA", end_color="D4EDDA", fill_type="solid")
    ws.merge_cells(start_row=total_row, start_column=1, end_row=total_row, end_column=4)
    cell_label = ws.cell(row=total_row, column=1, value="TOTAL")
    cell_label.font = total_font
    cell_label.alignment = Alignment(horizontal="right")
    cell_label.fill = total_fill
    cell_label.border = border
    for c in range(2, 5):
        ws.cell(row=total_row, column=c).fill = total_fill
        ws.cell(row=total_row, column=c).border = border
    cell_total = ws.cell(row=total_row, column=5, value=total)
    cell_total.font = total_font
    cell_total.number_format = '#,##0.00'
    cell_total.alignment = Alignment(horizontal="right")
    cell_total.fill = total_fill
    cell_total.border = border

    # Ajustar anchos
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 25
    ws.column_dimensions["C"].width = 30
    ws.column_dimensions["D"].width = 20
    ws.column_dimensions["E"].width = 15

    # Generar archivo
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"ventas_{periodo}_{valor}.xlsx"
    return send_file(output, download_name=filename, as_attachment=True, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


if __name__ == "__main__":
    init_db()
    MY_USER_ID, MY_USER_NAME, MY_USER_EMAIL = fetch_my_user_info()
    # Iniciar polling en hilo de fondo
    poll_thread = threading.Thread(target=poll_payments, daemon=True)
    poll_thread.start()
    print(f"Servidor iniciado en http://localhost:{FLASK_PORT}")
    print("Webhook + Polling cada 60s activos. Esperando pagos...")
    app.run(host="0.0.0.0", port=FLASK_PORT, debug=False)
